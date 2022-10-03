# coding:utf-8
# ok_net.py
from multiprocessing.resource_sharer import stop
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome import service as fs
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from time import sleep
import openpyxl
import os
import signal

# ログイン情報
myid = "*******"
mypass = "*******"

# エクセル設定
wb = openpyxl.load_workbook('./ok_list.xlsm')
ws = wb['Sheet1']
maxRow = ws.max_row
i = 0

# ブラウザを開く。
CHROMEDRIVER = "/Users/shunsukekumagai/dev/GitHub/ok_net/chromedriver"
chrome_service = fs.Service(executable_path=CHROMEDRIVER)
driver = webdriver.Chrome(service=chrome_service)

# Googleの検索TOP画面を開く。
driver.get("https://ok-netsuper.com/login")

# ログイン操作
mail = driver.find_element(By.ID, "username")
password = driver.find_element(By.ID, "password")
mail.send_keys(myid)
password.send_keys(mypass)
# 認証と配送時間を人が操作する時間待機
# wait = WebDriverWait(driver, timeout=30)
# wait.until(EC.presence_of_element_located(By.NAME, "header-bottom-loginuser-schedule"))
sleep(20)
driver.get("https://ok-netsuper.com/delivery-timetable/delivery-timetable-select")
driver.find_element(By.XPATH,"/html/body/div[2]/div[1]/div[2]/div/form/table/tbody/tr[3]/td[2]").click()
driver.find_element(By.ID,"continue-shopping").click()
sleep(2)

# タブを開く
driver.execute_script("window.open();")
# 開いたタブをアクティブにする
driver.switch_to.window(driver.window_handles[1])
# okホーム開く
driver.get("https://ok-netsuper.com/")



try:
# ここからリストがなくなるまでloop
    for i in range(2,maxRow + 1):

        if ws.cell(row = i, column = 2).value is not None:
            # リストから商品名を取得
            product_name = ws.cell(row = i, column = 1).value
            # リストから個数を取得
            value = ws.cell(row = i, column = 2).value
            # 商品を検索欄選択
            search_product = driver.find_element(By.CSS_SELECTOR, "input.header-bottom-search-input")
            # 検索欄をクリア
            search_product.clear()
            # 検索欄に商品名を記入
            search_product.send_keys(product_name)
            # ENTERを押す
            search_product.send_keys(Keys.ENTER)
            sleep(1)
            try:
                for j in range(value):
                    driver.find_elements(By.CSS_SELECTOR,".ok-button")[4].click()
                    sleep(1)
            except IndexError:
                # none_link = none_link + " " + product_name
                print(product_name)
                # ws.cell(row = i, column = 3).value = "検索なし"
                # wb.save(wb)
finally:
    os.kill(driver.service.process.pid,signal.SIGTERM)